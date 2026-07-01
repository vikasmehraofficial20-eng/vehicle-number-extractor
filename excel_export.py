from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def build_excel(results, video_filename, out_path, city='', garage='', auditor='', audit_date=''):
    wb = Workbook()
    ws = wb.active
    ws.title = "Detected Vehicles"

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
    normal_font = Font(name='Arial', size=11)
    low_conf_fill = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws['A1'] = 'Vehicle Number Detection Report'
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws.merge_cells('A1:I1')

    # Incoming date is an HTML date input value (YYYY-MM-DD); display it more readably
    date_display = audit_date or '-'
    if audit_date:
        try:
            date_display = datetime.strptime(audit_date, '%Y-%m-%d').strftime('%d-%b-%Y')
        except ValueError:
            pass

    ws['A2'] = (f'City: {city or "-"}    |    Garage/Location: {garage or "-"}    |    '
                f'Auditor: {auditor or "-"}    |    Date: {date_display}')
    ws['A2'].font = Font(name='Arial', bold=True, size=10.5, color='1F4E78')
    ws.merge_cells('A2:I2')

    ws['A3'] = f'Source video: {video_filename}    |    Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A3'].font = Font(name='Arial', italic=True, size=10, color='555555')
    ws.merge_cells('A3:I3')

    header_row = 5
    headers = ['S.No', 'City', 'Garage/Location', 'Auditor Name', 'Date', 'Vehicle Number',
               'Confidence (%)', 'Times Detected', 'First Seen (sec)']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    row = header_row + 1
    for i, r in enumerate(results, start=1):
        low_conf = r['confidence'] < 55
        vals = [i, city, garage, auditor, date_display, r['plate_number'],
                r['confidence'], r['frames_detected'], r['first_seen_seconds']]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = normal_font
            c.border = border
            c.alignment = Alignment(horizontal='center' if col not in (2, 3, 4, 6) else 'left',
                                     vertical='center')
            if low_conf:
                c.fill = low_conf_fill
        row += 1

    if not results:
        ws.cell(row=row, column=1, value='No plates detected.').font = Font(name='Arial', italic=True)
        row += 1

    note_row = row + 1
    ws.cell(row=note_row, column=1,
            value='Note: Rows highlighted in yellow have lower OCR confidence (<55%) — please verify these manually against the video.').font = \
        Font(name='Arial', italic=True, size=9, color='806000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)

    widths = [6, 16, 20, 18, 14, 22, 16, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'A{header_row+1}'

    wb.save(out_path)
    return out_path
